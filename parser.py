import openpyxl
import re

# Создаем новый Excel-файл или открываем существующий
wb = openpyxl.Workbook()
sheet = wb.active

# Задаем имя столбцам
sheet.cell(row=1, column=1).value = "Дата USD/RUB"
sheet.cell(row=1, column=2).value = "Курс USD/RUB"
sheet.cell(row=1, column=3).value = "Время USD/RUB"
sheet.cell(row=1, column=4).value = "Дата торгов JPY"
sheet.cell(row=1, column=5).value = "Курс JPY/RUB"
sheet.cell(row=1, column=6).value = "Время JPY/RUB"
sheet.cell(row=1, column=7).value = "Результат"


# Парсим данные из первого файла
row_num = 2
with open("./data1.xml", 'r') as file:
    for line in file:
        match = re.match(r'<row\s.*\/>', line)
        if match:
            trade_date = re.search(r'tradedate=\"(.+?)\"', line)
            trade_time = re.search(r'tradetime=\"(.+?)\"', line)
            rate = re.search(r'rate=\"(.+?)\"', line)
            # Записываем данные в Excel-файл
            sheet.cell(row=row_num, column=1).value = trade_date.group(1)
            sheet.cell(row=row_num, column=2).value = rate.group(1)
            sheet.cell(row=row_num, column=2).number_format = "#,##0.00₽"
            sheet.cell(row=row_num, column=3).value = trade_time.group(1)
            row_num += 1

# Парсим данные из второго файла
row_num = 2
with open("./data2.xml", 'r') as file:
    for line in file:
        match = re.match(r'<row\s.*\/>', line)
        if match:
            trade_date = re.search(r'tradedate=\"(.+?)\"', line)
            trade_time = re.search(r'tradetime=\"(.+?)\"', line)
            rate = re.search(r'rate=\"(.+?)\"', line)
            # Записываем данные в Excel-файл
            sheet.cell(row=row_num, column=4).value = trade_date.group(1)
            sheet.cell(row=row_num, column=5).value = rate.group(1)
            sheet.cell(row=row_num, column=5).number_format = "#,##0.00₽"
            sheet.cell(row=row_num, column=6).value = trade_time.group(1)
            row_num += 1

# Вычисляем результат и записываем его в Excel-файл
row_num = 2
for row in range(2, sheet.max_row + 1):
    rate_usd = float(sheet.cell(row=row, column=2).value)
    rate_jpy = float(sheet.cell(row=row, column=5).value)
    result = rate_usd / rate_jpy
    sheet.cell(row=row, column=7).value = result
    sheet.cell(row=row, column=7).number_format = "#,##0.00₽"
    row_num += 1


# Задаем автоширину столбцов
for column_cells in sheet.columns:
    length = max(len(str(cell.value)) for cell in column_cells)
    sheet.column_dimensions[column_cells[0].column_letter].width = length + 2

# Сохраняем Excel-файл
wb.save("report.xlsx")

# Отправка отчета на электронную почту через smtplib